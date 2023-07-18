import csv
from rest_framework import generics, status
from rest_framework.parsers import FileUploadParser
from rest_framework.exceptions import ValidationError
from io import TextIOWrapper
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from rest_framework.response import Response
from django.core.mail import send_mail
from django.conf import settings
from .serializers import *
from .models import *

# Create your views here.
class StudentCreateView(generics.CreateAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

class StudentListView(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer


class BulkImportView(generics.CreateAPIView):
    parser_class = (FileUploadParser,)

    def create(self, request, *args, **kwargs):
        file = request.data['file']
        csv_file = TextIOWrapper(file, encoding=request.encoding)
        reader = csv.DictReader(csv_file)
        students = []
        for row in reader:
            serializer = StudentSerializer(data=row)
            if serializer.is_valid():
                students.append(serializer.validated_data)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        Student.objects.bulk_create([Student(**data) for data in students])
        return Response("Data imported successfully", status=status.HTTP_201_CREATED)
    

class ParentCreateView(generics.CreateAPIView):
    queryset = Parent.objects.all()
    serializer_class = ParentSerializer

class ParentListView(generics.ListAPIView):
    queryset = Parent.objects.all()
    serializer_class = ParentSerializer

class AcademicDetailsCreateView(generics.CreateAPIView):
    queryset = Academic_Detail.objects.all()
    serializer_class = AcademicDetailsSerializer

class AcademicDetailsListView(generics.ListAPIView):
    queryset = Academic_Detail.objects.all()
    serializer_class = AcademicDetailsSerializer

class DocumentUploadCreateView(generics.CreateAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentUploadSerializer

class DocumentListView(generics.ListAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentUploadSerializer


class StudentFilterView(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    
    def get_queryset(self):
        #queryset define the parent class
        queryset = super().get_queryset()
        
        #filter by class, section, admission_category as arguments
        class_param = self.request.query_params.get('class_name', None)
        section_param = self.request.query_params.get('section', None)
        admission_category_param = self.request.query_params.get('admission_category', None)
        
        if class_param:
            queryset = queryset.filter(academic_detail__class_name=class_param)
        if section_param:
            queryset = queryset.filter(academic_detail__section=section_param)
        if admission_category_param:
            queryset = queryset.filter(admission_category=admission_category_param)
        
        # Custom validation
        if not queryset.exists():
            raise ValidationError("No matching records found.")

        return queryset


class StudentExportExcelView(generics.GenericAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def export_to_excel(self, queryset):
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Workbook"

        # Write column headers
        headers = ['Student Name', 'Class', 'Section']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data rows
        for row_num, student in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=student.student_name)
            
            # Retrieve the related Academic_Detail instance
            academic_detail = student.academic_detail_set.first()
        
            # Check if academic_detail exists
            if academic_detail:
                ws.cell(row=row_num, column=2, value=academic_detail.class_name)
                ws.cell(row=row_num, column=3, value=academic_detail.section)

        # Create the response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)

        return response

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Export all students to Excel
        return self.export_to_excel(queryset)


class StudentExportPDFView(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    def export_to_pdf(self, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=students.pdf'

        # Create a PDF canvas
        p = canvas.Canvas(response)

        # Write data to the PDF
        y = 750  # Starting y-coordinate for the content
        for student in queryset:
            p.drawString(50, y, f"Student Name: {student.student_name}")

            # Retrieve the related Academic_Detail instance
            academic_detail = student.academic_detail_set.first()

            # Check if academic_detail exists
            if academic_detail:
                p.drawString(50, y - 20, f"Class: {academic_detail.class_name}")
                p.drawString(50, y - 40, f"Section: {academic_detail.section}")

            y -= 80

        # Save the PDF and return the response
        p.showPage()
        p.save()
        return response

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Export all students to PDF
        return self.export_to_pdf(queryset)


class FinalSubmitView(generics.CreateAPIView):
    queryset = Academic_Detail.objects.all()
    serializer_class = AcademicDetailsSerializer

    def perform_create(self, serializer):
        # Save the academic details
        academic_detail = serializer.save()

        # Send email to the student
        student_subject = "Enrollment Confirmation - Dummy School"
        student_message = f"Dear {academic_detail.student.student_name},\n\nYou have been enrolled in Dummy School. Your Enrollment ID is {academic_detail.enrollment_id}. Please provide us with the necessary hard documents for future reference.\n\nTeam\nDummy School"
        send_mail(student_subject, student_message, settings.DEFAULT_FROM_EMAIL, [academic_detail.student.email])

        # Send email notification to the admin
        admin_subject = "New Student Enrollment Notification - Dummy School"
        admin_message = f"Dear Admin,\n\nA new student, {academic_detail.student.student_name}, has been enrolled in class {academic_detail.class_name}, section {academic_detail.section} with Enrollment ID {academic_detail.enrollment_id} for the {academic_detail.date_of_joining} session."
        send_mail(admin_subject, admin_message, settings.DEFAULT_FROM_EMAIL, ['durganand.jha@habrie.com'])

        return Response(serializer.data, status=201)